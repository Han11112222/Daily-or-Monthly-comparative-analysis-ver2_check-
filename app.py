import calendar
from io import BytesIO
from pathlib import Path

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


# ─────────────────────────────────────────────
# 1. 단위/환산 상수
# ─────────────────────────────────────────────
MJ_PER_NM3 = 42.563
MJ_TO_GJ = 1.0 / 1000.0

def mj_to_gj(x):
    try: return x * MJ_TO_GJ
    except Exception: return np.nan

def mj_to_m3(x):
    try: return x / MJ_PER_NM3
    except Exception: return np.nan


# ─────────────────────────────────────────────
# 2. 기본 설정 & 세션 상태 초기화
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="도시가스 공급량: 일별계획 예측 (Final)",
    layout="wide",
)

if 'rec_active' not in st.session_state: st.session_state['rec_active'] = False
if 'prev_active' not in st.session_state: st.session_state['prev_active'] = False

if 'cal_start' not in st.session_state: st.session_state['cal_start'] = None
if 'cal_end' not in st.session_state: st.session_state['cal_end'] = None
if 'fix_start' not in st.session_state: st.session_state['fix_start'] = None
if 'fix_end' not in st.session_state: st.session_state['fix_end'] = None
if 'rec_rate' not in st.session_state: st.session_state['rec_rate'] = 0.0


# ─────────────────────────────────────────────
# 3. 데이터 불러오기
# ─────────────────────────────────────────────
@st.cache_data
def load_daily_data():
    excel_path = Path(__file__).parent / "공급량(일일실적).xlsx"
    if not excel_path.exists():
        return pd.DataFrame(), pd.DataFrame()

    try:
        df_raw = pd.read_excel(excel_path)
        # 필요한 컬럼만 추출 및 이름 통일
        required = ["일자", "공급량(MJ)", "공급량(M3)", "평균기온(℃)"]
        cols = [c for c in required if c in df_raw.columns]
        df_raw = df_raw[cols].copy()
        
        df_raw["일자"] = pd.to_datetime(df_raw["일자"])
        df_raw["연도"] = df_raw["일자"].dt.year
        df_raw["월"] = df_raw["일자"].dt.month
        df_raw["일"] = df_raw["일자"].dt.day
        
        # 미리 계산
        df_raw["weekday_idx"] = df_raw["일자"].dt.weekday
        df_raw["nth_dow"] = df_raw.groupby(["연도", "월", "weekday_idx"]).cumcount() + 1

        df_temp_all = df_raw.dropna(subset=["평균기온(℃)"]).copy() if "평균기온(℃)" in df_raw.columns else df_raw
        df_model = df_raw.dropna(subset=["공급량(MJ)"]).copy() if "공급량(MJ)" in df_raw.columns else df_raw
        return df_model, df_temp_all
    except:
        return pd.DataFrame(), pd.DataFrame()

@st.cache_data
def load_monthly_plan() -> pd.DataFrame:
    excel_path = Path(__file__).parent / "공급량(계획_실적).xlsx"
    if not excel_path.exists(): return pd.DataFrame()
    try:
        df = pd.read_excel(excel_path, sheet_name="월별계획_실적")
        if "연" in df.columns: df["연"] = pd.to_numeric(df["연"], errors='coerce').fillna(0).astype(int)
        if "월" in df.columns: df["월"] = pd.to_numeric(df["월"], errors='coerce').fillna(0).astype(int)
        return df
    except: return pd.DataFrame()

@st.cache_data
def load_effective_calendar() -> pd.DataFrame | None:
    excel_path = Path(__file__).parent / "effective_days_calendar.xlsx"
    if not excel_path.exists(): return None
    try:
        df = pd.read_excel(excel_path)
        if "날짜" not in df.columns: return None
        df["일자"] = pd.to_datetime(df["날짜"].astype(str), format="%Y%m%d", errors="coerce")
        for col in ["공휴일여부", "명절여부"]:
            if col not in df.columns: df[col] = False
            if df[col].dtype != bool:
                df[col] = df[col].fillna(False).astype(bool)
        return df[["일자", "공휴일여부", "명절여부"]].copy()
    except: return None


# ─────────────────────────────────────────────
# 4. 유틸 함수들
# ─────────────────────────────────────────────
def _find_plan_col(df_plan: pd.DataFrame) -> str:
    candidates = ["계획(사업계획제출_MJ)", "계획(사업계획제출)", "계획_MJ", "계획"]
    for c in candidates:
        if c in df_plan.columns: return c
    nums = [c for c in df_plan.columns if pd.api.types.is_numeric_dtype(df_plan[c]) and c not in ["연", "월"]]
    return nums[0] if nums else "계획(사업계획제출_MJ)"

def make_month_plan_horizontal(df_plan: pd.DataFrame, target_year: int, plan_col: str) -> pd.DataFrame:
    if df_plan.empty or not plan_col: return pd.DataFrame()
    df_year = df_plan[df_plan["연"] == target_year][["월", plan_col]].copy()
    if df_year.empty: return pd.DataFrame()
    base = pd.DataFrame({"월": list(range(1, 13))})
    df_year = base.merge(df_year, on="월", how="left")
    df_year = df_year.rename(columns={plan_col: "월별 계획(MJ)"})
    
    total_mj = df_year["월별 계획(MJ)"].sum()
    df_year["월별 계획(GJ)"] = (df_year["월별 계획(MJ)"].apply(mj_to_gj)).round(0)
    df_year["월별 계획(㎥)"] = (df_year["월별 계획(MJ)"].apply(mj_to_m3)).round(0)
    
    total_gj = mj_to_gj(total_mj)
    total_m3 = mj_to_m3(total_mj)
    
    row_gj = {"구분": "사업계획(월별 계획, GJ)"}
    row_m3 = {"구분": "사업계획(월별 계획, ㎥)"}
    
    for _, row in df_year.iterrows():
        m = int(row["월"])
        mj = row["월별 계획(MJ)"]
        row_gj[f"{m}월"] = round(mj_to_gj(mj), 0)
        row_m3[f"{m}월"] = round(mj_to_m3(mj), 0)
    
    row_gj["연간합계"] = round(total_gj, 0)
    row_m3["연간합계"] = round(total_m3, 0)
    
    return pd.DataFrame([row_gj, row_m3])

def format_table_generic(df, percent_cols=None):
    if df.empty: return df
    df = df.copy()
    percent_cols = percent_cols or []
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime('%Y-%m-%d')
        elif df[col].dtype == bool:
            df[col] = df[col].map(lambda x: "O" if x else "")
        elif col == "Diff(%)": 
            df[col] = df[col].map(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
        elif col in percent_cols:
            df[col] = df[col].map(lambda x: f"{x:.4f}" if pd.notna(x) else "")
        elif pd.api.types.is_numeric_dtype(df[col]):
             if col in ["연", "월", "일", "WeekNum"]:
                 df[col] = df[col].map(lambda x: f"{int(x)}" if pd.notna(x) else "")
             else:
                 df[col] = df[col].map(lambda x: f"{x:,.0f}" if pd.notna(x) else "")
    return df

def show_table_no_index(df: pd.DataFrame, height: int = 260):
    try: st.dataframe(df, use_container_width=True, hide_index=True, height=height)
    except: st.table(df)

def _format_excel_sheet(ws, freeze="A2", center=True):
    if freeze: ws.freeze_panes = freeze
    if center:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for c in row: c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _add_cumulative_status_sheet(wb, annual_year: int):
    sheet_name = "누적계획현황"
    if sheet_name in wb.sheetnames: return
    ws = wb.create_sheet(sheet_name)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    ws["A1"] = "기준일"; ws["B1"] = pd.Timestamp(f"{annual_year}-01-01")
    ws["B1"].number_format = "yyyy-mm-dd"
    
    headers = ["구분", "목표(GJ)", "누적(GJ)", "목표(m³)", "누적(m³)", "진행률(GJ)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=3, column=j+1, value=h).border = border
    ws.freeze_panes = "A4"

def _make_display_table_gj_m3(df_mj: pd.DataFrame) -> pd.DataFrame:
    df = df_mj.copy()
    for base_col in ["최근N년_평균공급량(MJ)", "최근N년_총공급량(MJ)", "예상공급량(MJ)", "보정_예상공급량(MJ)"]:
        if base_col not in df.columns: continue
        gj_col = base_col.replace("(MJ)", "(GJ)")
        m3_col = base_col.replace("(MJ)", "(㎥)")
        df[gj_col] = df[base_col].apply(mj_to_gj).round(0)
        df[m3_col] = df[base_col].apply(mj_to_m3).round(0)
    
    df_disp = df.rename(columns={
        "예상공급량(GJ)": "As-Is(기존)",
        "보정_예상공급량(GJ)": "To-Be(보정)"
    })
    if "To-Be(보정)" in df_disp.columns and "As-Is(기존)" in df_disp.columns:
        df_disp["Diff(증감)"] = df_disp["To-Be(보정)"] - df_disp["As-Is(기존)"]
        df_disp["Diff(%)"] = df_disp.apply(
            lambda row: (row["Diff(증감)"] / row["As-Is(기존)"] * 100) if row["As-Is(기존)"] != 0 else 0, axis=1
        )
        
    keep = ["일자", "요일", "구분", "일별비율", "As-Is(기존)", "To-Be(보정)", "Diff(증감)", "Diff(%)", "is_outlier"]
    final_cols = [c for c in keep if c in df_disp.columns]
    return df_disp[final_cols].copy()


# ─────────────────────────────────────────────
# 5. 핵심 분석 로직 (Daily)
# ─────────────────────────────────────────────
def make_daily_plan_table(df_daily, df_plan, target_year, target_month, recent_window, apply_trend=False):
    trend_msg = ""
    
    cal_df = load_effective_calendar()
    plan_col = _find_plan_col(df_plan)
    
    all_years = sorted(df_daily["연도"].unique())
    start_year = target_year - recent_window
    candidate_years = [y for y in range(start_year, target_year) if y in all_years]
    
    if len(candidate_years) == 0: return None, None, [], pd.DataFrame(), ""
    
    df_pool = df_daily[(df_daily["연도"].isin(candidate_years)) & (df_daily["월"] == target_month)].copy()
    df_pool = df_pool.dropna(subset=["공급량(MJ)"])
    used_years = sorted(df_pool["연도"].unique().tolist())
    if not used_years: return None, None, [], pd.DataFrame(), ""

    df_recent = df_daily[(df_daily["연도"].isin(used_years)) & (df_daily["월"] == target_month)].copy()
    df_recent = df_recent.dropna(subset=["공급량(MJ)"])
    if df_recent.empty: return None, None, used_years, pd.DataFrame(), ""

    # [형님 로직] 데이터 전처리
    df_recent = df_recent.sort_values(["연도", "일"]).copy()
    df_recent["weekday_idx"] = df_recent["일자"].dt.weekday

    if cal_df is not None:
        df_recent = df_recent.merge(cal_df, on="일자", how="left")
        if ("공휴일여부" not in df_recent.columns) and ("공휴일여버" in df_recent.columns):
            df_recent = df_recent.rename(columns={"공휴일여버": "공휴일여부"})
        if "공휴일여부" not in df_recent.columns:
            df_recent["공휴일여부"] = False

        df_recent["공휴일여부"] = df_recent["공휴일여부"].fillna(False).astype(bool)
        df_recent["명절여부"] = df_recent["명절여부"].fillna(False).astype(bool)
    else:
        df_recent["공휴일여부"] = False
        df_recent["명절여부"] = False

    df_recent["is_holiday"] = df_recent["공휴일여부"] | df_recent["명절여부"]
    df_recent["is_weekend"] = (df_recent["weekday_idx"] >= 5) | df_recent["is_holiday"]
    df_recent["is_weekday1"] = (~df_recent["is_weekend"]) & (df_recent["weekday_idx"].isin([0, 4]))
    df_recent["is_weekday2"] = (~df_recent["is_weekend"]) & (df_recent["weekday_idx"].isin([1, 2, 3]))

    # [형님 로직] 비율 계산
    df_recent["month_total"] = df_recent.groupby("연도")["공급량(MJ)"].transform("sum")
    df_recent["ratio"] = df_recent["공급량(MJ)"] / df_recent["month_total"]

    df_recent["nth_dow"] = (
        df_recent.sort_values(["연도", "일"])
        .groupby(["연도", "weekday_idx"])
        .cumcount()
        + 1
    )

    weekend_mask = df_recent["is_weekend"]
    w1_mask = df_recent["is_weekday1"]
    w2_mask = df_recent["is_weekday2"]

    ratio_weekend_group = (
        df_recent[weekend_mask].groupby(["weekday_idx", "nth_dow"])["ratio"].mean()
        if df_recent[weekend_mask].size > 0 else pd.Series(dtype=float)
    )
    ratio_w1_group = (
        df_recent[w1_mask].groupby(["weekday_idx", "nth_dow"])["ratio"].mean()
        if df_recent[w1_mask].size > 0 else pd.Series(dtype=float)
    )
    ratio_w2_group = (
        df_recent[w2_mask].groupby(["weekday_idx", "nth_dow"])["ratio"].mean()
        if df_recent[w2_mask].size > 0 else pd.Series(dtype=float)
    )

    ratio_weekend_by_dow = (
        df_recent[weekend_mask].groupby("weekday_idx")["ratio"].mean()
        if df_recent[weekend_mask].size > 0 else pd.Series(dtype=float)
    )
    ratio_w1_by_dow = (
        df_recent[w1_mask].groupby("weekday_idx")["ratio"].mean()
        if df_recent[w1_mask].size > 0 else pd.Series(dtype=float)
    )
    ratio_w2_by_dow = (
        df_recent[w2_mask].groupby("weekday_idx")["ratio"].mean()
        if df_recent[w2_mask].size > 0 else pd.Series(dtype=float)
    )

    ratio_weekend_group_dict = ratio_weekend_group.to_dict()
    ratio_weekend_by_dow_dict = ratio_weekend_by_dow.to_dict()
    ratio_w1_group_dict = ratio_w1_group.to_dict()
    ratio_w1_by_dow_dict = ratio_w1_by_dow.to_dict()
    ratio_w2_group_dict = ratio_w2_group.to_dict()
    ratio_w2_by_dow_dict = ratio_w2_by_dow.to_dict()

    last_day = calendar.monthrange(target_year, target_month)[1]
    date_range = pd.date_range(f"{target_year}-{target_month:02d}-01", periods=last_day, freq="D")

    df_target = pd.DataFrame({"일자": date_range})
    df_target["연"] = target_year
    df_target["월"] = target_month
    df_target["일"] = df_target["일자"].dt.day
    df_target["weekday_idx"] = df_target["일자"].dt.weekday

    if cal_df is not None:
        df_target = df_target.merge(cal_df, on="일자", how="left")
        if ("공휴일여부" not in df_target.columns) and ("공휴일여버" in df_target.columns):
            df_target = df_target.rename(columns={"공휴일여버": "공휴일여부"})
        if "공휴일여부" not in df_target.columns:
            df_target["공휴일여부"] = False

        df_target["공휴일여부"] = df_target["공휴일여부"].fillna(False).astype(bool)
        df_target["명절여부"] = df_target["명절여부"].fillna(False).astype(bool)
    else:
        df_target["공휴일여부"] = False
        df_target["명절여부"] = False

    df_target["is_holiday"] = df_target["공휴일여부"] | df_target["명절여부"]
    df_target["is_weekend"] = (df_target["weekday_idx"] >= 5) | df_target["is_holiday"]
    df_target["is_weekday1"] = (~df_target["is_weekend"]) & (df_target["weekday_idx"].isin([0, 4]))
    df_target["is_weekday2"] = (~df_target["is_weekend"]) & (df_target["weekday_idx"].isin([1, 2, 3]))

    weekday_names = ["월", "화", "수", "목", "금", "토", "일"]
    df_target["요일"] = df_target["weekday_idx"].map(lambda i: weekday_names[i])

    df_target["nth_dow"] = df_target.sort_values("일").groupby("weekday_idx").cumcount() + 1

    def _label(row):
        if row["is_weekend"]: return "주말/공휴일"
        if row["is_weekday1"]: return "평일1(월,금)"
        return "평일2(화,수,목)"
    df_target["구분"] = df_target.apply(_label, axis=1)

    # [형님 로직] 비율 적용
    def _pick_ratio(row):
        dow = int(row["weekday_idx"])
        nth = int(row["nth_dow"])
        key = (dow, nth)

        if bool(row["is_weekend"]):
            v = ratio_weekend_group_dict.get(key, None)
            if v is None or pd.isna(v):
                v = ratio_weekend_by_dow_dict.get(dow, None)
            return v

        if bool(row["is_weekday1"]):
            v = ratio_w1_group_dict.get(key, None)
            if v is None or pd.isna(v):
                v = ratio_w1_by_dow_dict.get(dow, None)
            return v

        v = ratio_w2_group_dict.get(key, None)
        if v is None or pd.isna(v):
            v = ratio_w2_by_dow_dict.get(dow, None)
        return v

    df_target["raw"] = df_target.apply(_pick_ratio, axis=1).astype("float64")

    overall_mean = df_target["raw"].dropna().mean() if df_target["raw"].notna().any() else np.nan
    for cat in ["주말/공휴일", "평일1(월,금)", "평일2(화,수,목)"]:
        mask = df_target["구분"] == cat
        if mask.any():
            m = df_target.loc[mask, "raw"].dropna().mean()
            if pd.isna(m): m = overall_mean
            df_target.loc[mask, "raw"] = df_target.loc[mask, "raw"].fillna(m)

    if df_target["raw"].isna().all(): df_target["raw"] = 1.0

    raw_sum = df_target["raw"].sum()
    df_target["일별비율"] = (df_target["raw"] / raw_sum) if raw_sum > 0 else (1.0 / last_day)

    month_total_all = df_recent["공급량(MJ)"].sum()
    df_target["최근N년_총공급량(MJ)"] = df_target["일별비율"] * month_total_all
    df_target["최근N년_평균공급량(MJ)"] = df_target["최근N년_총공급량(MJ)"] / len(used_years)

    row_plan = df_plan[(df_plan["연"] == target_year) & (df_plan["월"] == target_month)]
    plan_total = float(row_plan[plan_col].iloc[0]) if not row_plan.empty else 0
    
    df_target["예상공급량(MJ)"] = (df_target["일별비율"] * plan_total).round(0)
    
    # [Bound 추가]
    df_target["WeekNum"] = df_target["일자"].dt.isocalendar().week
    df_target["Group_Mean"] = df_target.groupby(["WeekNum", "is_weekend"])["예상공급량(MJ)"].transform("mean")
    df_target["Bound_Upper"] = df_target["Group_Mean"] * 1.10
    df_target["Bound_Lower"] = df_target["Group_Mean"] * 0.90
    df_target["is_outlier"] = (df_target["예상공급량(MJ)"] > df_target["Bound_Upper"]) | (df_target["예상공급량(MJ)"] < df_target["Bound_Lower"])

    df_mat = df_recent.pivot_table(index="일", columns="연도", values="공급량(MJ)", aggfunc="sum").sort_index(axis=1)
    df_debug = df_target.copy()

    return df_target, df_mat, used_years, df_debug, trend_msg

def _build_year_daily_plan(df_daily, df_plan, target_year, recent_window):
    all_rows = []
    month_summary_rows = []
    plan_col = _find_plan_col(df_plan)
    
    for m in range(1, 13):
        res, _, _, _, _ = make_daily_plan_table(df_daily, df_plan, target_year, m, recent_window, apply_trend=False)
        row_plan = df_plan[(df_plan["연"] == target_year) & (df_plan["월"] == m)]
        plan_total_mj = float(row_plan[plan_col].iloc[0]) if not row_plan.empty else np.nan
        
        if res is not None:
             all_rows.append(res)
             month_summary_rows.append({
                "월": m,
                "월간 계획(GJ)": round(mj_to_gj(plan_total_mj), 0),
                "월간 계획(㎥)": round(mj_to_m3(plan_total_mj), 0)
             })

    if not all_rows: return pd.DataFrame(), pd.DataFrame()
    df_year = pd.concat(all_rows, ignore_index=True)
    return df_year, pd.DataFrame(month_summary_rows)

# ─────────────────────────────────────────────
# 6. UI 및 시각화
# ─────────────────────────────────────────────
def tab_daily_plan(df_daily: pd.DataFrame):
    st.subheader("📅 Daily 공급량 분석 — 최근 N년 패턴 기반 일별 계획")

    uploaded_file = st.sidebar.file_uploader("📂 비교용 엑셀/CSV 파일 업로드", type=["xlsx", "csv"])

    df_plan = load_monthly_plan()
    plan_col = _find_plan_col(df_plan)
    years_plan = sorted(df_plan["연"].unique())
    default_year_idx = years_plan.index(2026) if 2026 in years_plan else len(years_plan) - 1

    col_y, col_m, _ = st.columns([1, 1, 2])
    with col_y: target_year = st.selectbox("계획 연도 선택", years_plan, index=default_year_idx)
    with col_m:
        months_plan = sorted(df_plan[df_plan["연"] == target_year]["월"].unique())
        default_month_idx = months_plan.index(1) if 1 in months_plan else 0
        target_month = st.selectbox("계획 월 선택", months_plan, index=default_month_idx, format_func=lambda m: f"{m}월")

    all_years = sorted(df_daily["연도"].unique())
    hist_years = [y for y in all_years if y < target_year]
    if len(hist_years) < 1: st.warning("데이터 부족"); return

    slider_min = 1; slider_max = min(10, len(hist_years))
    col_slider, _ = st.columns([2, 3])
    with col_slider:
        recent_window = st.slider("최근 몇 년 평균?", min_value=slider_min, max_value=slider_max, value=min(3, slider_max), step=1)

    apply_trend = st.checkbox("📉 추세적용 (월초 vs 월말 기온반영)", value=False)

    df_result, df_mat, used_years, df_debug, trend_msg = make_daily_plan_table(
        df_daily, df_plan, target_year, target_month, recent_window, apply_trend=apply_trend
    )

    if apply_trend and trend_msg:
        st.info(trend_msg)

    if df_result is None: st.warning("데이터 부족"); return
    
    st.markdown(f"- 실제 학습 연도: {min(used_years)} ~ {max(used_years)}")
    plan_total_gj = mj_to_gj(df_result["예상공급량(MJ)"].sum())
    st.markdown(f"**{target_year}년 {target_month}월 합계:** `{plan_total_gj:,.0f} GJ`")

    view = df_result.copy()
    view["보정_예상공급량(MJ)"] = view["예상공급량(MJ)"]
    view["예상공급량(GJ)"] = view["예상공급량(MJ)"].apply(mj_to_gj)
    view["보정_예상공급량(GJ)"] = view["보정_예상공급량(MJ)"].apply(mj_to_gj)
    
    view["WeekNum"] = view["일자"].dt.isocalendar().week
    view["Group_Mean"] = view.groupby(["WeekNum", "is_weekend"])["예상공급량(MJ)"].transform("mean")
    view["Bound_Upper"] = view["Group_Mean"] * 1.10
    view["Bound_Lower"] = view["Group_Mean"] * 0.90
    view["Bound_Upper(GJ)"] = view["Bound_Upper"].apply(mj_to_gj)
    view["Bound_Lower(GJ)"] = view["Bound_Lower"].apply(mj_to_gj)
    view["is_outlier"] = (view["예상공급량(MJ)"] > view["Bound_Upper"]) | (view["예상공급량(MJ)"] < view["Bound_Lower"])
    
    st.divider()
    
    chart_placeholder = st.empty()

    if uploaded_file is not None:
        try:
            file_bytes = uploaded_file.getvalue()
            df_up = None
            
            try: df_up = pd.read_excel(BytesIO(file_bytes))
            except: pass
            
            if df_up is None:
                for enc in ['utf-8', 'cp949', 'euc-kr']:
                    try: df_up = pd.read_csv(BytesIO(file_bytes), encoding=enc); break
                    except: pass

            if df_up is None:
                st.error("❌ 파일을 읽을 수 없습니다. (Excel/CSV 포맷 확인 요망)")
            else:
                df_up.columns = df_up.columns.str.strip()
                
                target_col = None
                as_is_col = None
                
                for c in df_up.columns:
                    if "To-Be" in c and "최종" in c: target_col = c
                    if "As-Is" in c: as_is_col = c
                
                if target_col and "일자" in df_up.columns:
                    df_up["일자"] = pd.to_datetime(df_up["일자"], errors='coerce')
                    df_up = df_up.dropna(subset=["일자"])
                    df_up["일자"] = df_up["일자"].dt.normalize()
                    
                    df_up = df_up[
                        (df_up["일자"].dt.year == target_year) & 
                        (df_up["일자"].dt.month == target_month)
                    ].copy()
                    
                    if df_up.empty:
                        st.warning(f"⚠️ 업로드된 파일에 {target_year}년 {target_month}월 데이터가 없습니다.")
                    else:
                        if df_up[target_col].dtype == object:
                            df_up[target_col] = pd.to_numeric(df_up[target_col].astype(str).str.replace(',', ''), errors='coerce')
                        if as_is_col and df_up[as_is_col].dtype == object:
                            df_up[as_is_col] = pd.to_numeric(df_up[as_is_col].astype(str).str.replace(',', ''), errors='coerce')

                        agg_dict = {target_col: 'mean'}
                        if as_is_col: agg_dict[as_is_col] = 'mean'
                        df_up = df_up.groupby("일자", as_index=False).agg(agg_dict)

                        if df_up[target_col].mean() > 2000000:
                            df_up[target_col] = df_up[target_col] * 0.001
                            if as_is_col: df_up[as_is_col] = df_up[as_is_col] * 0.001
                            st.toast("💡 업로드된 파일의 단위를 MJ → GJ로 자동 변환했습니다.")

                        view_base = view[["일자", "예상공급량(GJ)", "Bound_Upper", "Bound_Lower"]].copy()
                        view_base["일자"] = view_base["일자"].dt.normalize() 
                        
                        df_merged = view_base.merge(df_up, on="일자", how="left")
                        
                        final_as_is = "Final_As_Is"
                        if as_is_col:
                            df_merged[final_as_is] = df_merged[as_is_col].fillna(df_merged["예상공급량(GJ)"])
                            df_merged.loc[df_merged[final_as_is] == 0, final_as_is] = df_merged["예상공급량(GJ)"]
                        else:
                            df_merged[final_as_is] = df_merged["예상공급량(GJ)"]

                        df_merged["Bound_Upper(GJ)"] = df_merged["Bound_Upper"].apply(mj_to_gj)
                        df_merged["Bound_Lower(GJ)"] = df_merged["Bound_Lower"].apply(mj_to_gj)

                        df_merged["weekday_idx"] = df_merged["일자"].dt.weekday
                        df_merged["is_weekend"] = df_merged["weekday_idx"] >= 5
                        df_merged["is_weekday1"] = (~df_merged["is_weekend"]) & (df_merged["weekday_idx"].isin([0, 4]))
                        df_merged["is_weekday2"] = (~df_merged["is_weekend"]) & (df_merged["weekday_idx"].isin([1, 2, 3]))
                        
                        def _get_label_up(r):
                            if r["is_weekend"]: return "주말/공휴일"
                            if r["is_weekday1"]: return "평일1(월,금)"
                            return "평일2(화,수,목)"
                        df_merged["구분"] = df_merged.apply(_get_label_up, axis=1)
                        
                        fig_up = go.Figure()
                        
                        u1 = df_merged[df_merged["구분"] == "평일1(월,금)"]
                        u2 = df_merged[df_merged["구분"] == "평일2(화,수,목)"]
                        ue = df_merged[df_merged["구분"] == "주말/공휴일"]
                        
                        fig_up.add_trace(go.Bar(x=u1["일자"].dt.day, y=u1[final_as_is], name="As-Is: 평일1(월,금)", marker_color="#1F77B4", width=0.8, hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))
                        fig_up.add_trace(go.Bar(x=u2["일자"].dt.day, y=u2[final_as_is], name="As-Is: 평일2(화,수,목)", marker_color="#87CEFA", width=0.8, hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))
                        fig_up.add_trace(go.Bar(x=ue["일자"].dt.day, y=ue[final_as_is], name="As-Is: 주말/공휴일", marker_color="#D62728", width=0.8, hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))
                        
                        if target_col in df_merged.columns:
                            mask_changed = (abs(df_merged[final_as_is] - df_merged[target_col]) > 1)
                            target_view = df_merged[mask_changed]
                            
                            # [수정] To-Be 막대 테두리 추가 (진한 선)
                            fig_up.add_trace(go.Bar(
                                x=target_view["일자"].dt.day, 
                                y=target_view[target_col],
                                marker_color="rgba(100, 100, 100, 0.4)", 
                                marker_line_color="rgba(60, 60, 60, 1.0)",
                                marker_line_width=2,
                                name="To-Be(보정)",
                                width=0.8,
                                hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"
                            ))
                        
                        fig_up.add_trace(go.Scatter(x=df_merged["일자"].dt.day, y=df_merged["Bound_Upper(GJ)"], mode='lines', line=dict(width=0), showlegend=False, hoverinfo='skip'))
                        fig_up.add_trace(go.Scatter(x=df_merged["일자"].dt.day, y=df_merged["Bound_Lower(GJ)"], mode='lines', line=dict(width=0), fill='tonexty', fillcolor='rgba(100,100,100,0.45)', name='범위(±10%)', hoverinfo='skip'))

                        fig_up.update_layout(
                            title=f"📂 업로드 데이터 ({target_year}년 {target_month}월): {uploaded_file.name}",
                            xaxis_title="일",
                            yaxis=dict(title="공급량(GJ)"),
                            barmode="overlay",
                            legend=dict(orientation="h", y=1.1)
                        )
                        st.plotly_chart(fig_up, use_container_width=True)
                else:
                    st.warning("⚠️ 업로드된 파일에 '일자' 또는 'To-Be(보정)_최종' 컬럼이 없습니다.")
                
        except Exception as e:
            st.error(f"파일 처리 중 상세 오류: {e}")
    
    _, col_btn = st.columns([5, 1]) 
    with col_btn:
        use_calib = st.checkbox("✅ 이상치 보정 활성화", value=False)
        
    diff_mj = 0
    mask_out = pd.Series([False]*len(view))

    if use_calib:
        c_rec1, c_rec2 = st.columns(2)
        
        if st.session_state['rec_active']:
            if c_rec1.button("✅ 추천 보정 적용중 (해제)", type="primary"):
                st.session_state['rec_active'] = False
                st.rerun()
        else:
            if c_rec1.button("🚀 추천 보정"):
                st.session_state['rec_active'] = True
                min_date = view["일자"].min().date()
                max_date = view["일자"].max().date()
                outliers = view[view["is_outlier"]]
                if not outliers.empty:
                    max_row = outliers.loc[outliers["예상공급량(MJ)"].idxmax()]
                    st.session_state['cal_start'] = max_row["일자"].date()
                    st.session_state['cal_end'] = max_row["일자"].date()
                    dev = (max_row["Bound_Upper"] - max_row["예상공급량(MJ)"]) / max_row["예상공급량(MJ)"] * 100
                    st.session_state['rec_rate'] = float(round(dev, 1))
                    
                    view_clean = view[view["일자"].dt.date != max_row["일자"].date()]
                    if not view_clean.empty:
                        best_week = view_clean.groupby("WeekNum")["예상공급량(MJ)"].sum().idxmax()
                        week_rows = view_clean[view_clean["WeekNum"] == best_week]
                        st.session_state['fix_start'] = week_rows["일자"].min().date()
                        st.session_state['fix_end'] = week_rows["일자"].max().date()
                    else:
                        st.session_state['fix_start'] = min_date
                        st.session_state['fix_end'] = max_date
                st.rerun()

        with st.expander("🛠️ 보정 구간 및 재배분 설정", expanded=True):
            min_d = view["일자"].min().date(); max_d = view["일자"].max().date()
            
            def validate_date(d):
                if d is None: return min_d
                if d < min_d or d > max_d: return min_d
                return d

            def_start = validate_date(st.session_state.get('cal_start'))
            def_end = validate_date(st.session_state.get('cal_end'))
            def_fix_s = validate_date(st.session_state.get('fix_start'))
            def_fix_e = validate_date(st.session_state.get('fix_end'))
            def_rate = st.session_state.get('rec_rate', 0.0)

            if def_end < def_start: def_end = def_start
            if def_fix_e < def_fix_s: def_fix_e = def_fix_s

            c1, c2 = st.columns(2)
            d_out = c1.date_input("1. 이상구간 (Outlier)", (def_start, def_end), min_value=min_d, max_value=max_d)
            d_fix = c2.date_input("2. 보정 구간 (Redistribution)", (def_fix_s, def_fix_e), min_value=min_d, max_value=max_d)
            
            if st.session_state['prev_active']:
                if st.button("✅ 전년도 실적 적용중 (해제)", type="primary"):
                    st.session_state['prev_active'] = False
                    st.rerun()
            else:
                if st.button("📅 전년도 실적 적용 (요일/주차 패턴 매칭)"):
                    st.session_state['prev_active'] = True
                    st.rerun()

            cal_rate = st.number_input("조정 비율 (%)", min_value=-50.0, max_value=50.0, value=float(def_rate), step=1.0)
            do_smooth = st.checkbox("🌊 평탄화 적용")

            if st.session_state['prev_active']:
                if isinstance(d_fix, tuple) and len(d_fix) == 2:
                    s_f, e_f = d_fix
                    target_mask = (view["일자"].dt.date >= s_f) & (view["일자"].dt.date <= e_f)
                    prev_year = target_year - 1
                    for idx, row in view[target_mask].iterrows():
                        cur_month = row["월"]
                        cur_wd = row["weekday_idx"]
                        cur_nth = row["nth_dow"]
                        
                        match = df_daily[
                            (df_daily["연도"] == prev_year) &
                            (df_daily["월"] == cur_month) &
                            (df_daily["weekday_idx"] == cur_wd) &
                            (df_daily["nth_dow"] == cur_nth)
                        ]
                        
                        if not match.empty:
                            view.loc[idx, "보정_예상공급량(MJ)"] = match.iloc[0]["공급량(MJ)"]
                        else:
                            fallback = row["일자"] - pd.Timedelta(weeks=52)
                            match_fb = df_daily[df_daily["일자"] == fallback]
                            if not match_fb.empty:
                                view.loc[idx, "보정_예상공급량(MJ)"] = match_fb.iloc[0]["공급량(MJ)"]

            if isinstance(d_out, tuple) and len(d_out) == 2 and isinstance(d_fix, tuple) and len(d_fix) == 2:
                s_out, e_out = d_out; s_fix, e_fix = d_fix
                
                mask_out = (view["일자"].dt.date >= s_out) & (view["일자"].dt.date <= e_out)
                mask_fix = (view["일자"].dt.date >= s_fix) & (view["일자"].dt.date <= e_fix)
                mask_fix = mask_fix & (~mask_out)

                if mask_out.any():
                    view.loc[mask_out, "보정_예상공급량(MJ)"] = view.loc[mask_out, "예상공급량(MJ)"] * (1 + cal_rate / 100.0)
                    diff_mj = (view.loc[mask_out, "예상공급량(MJ)"] - view.loc[mask_out, "보정_예상공급량(MJ)"]).sum()
                    
                    sum_r = view.loc[mask_fix, "일별비율"].sum()
                    if mask_fix.any() and sum_r > 0:
                        view.loc[mask_fix, "보정_예상공급량(MJ)"] += diff_mj * (view.loc[mask_fix, "일별비율"] / sum_r)
                        if do_smooth:
                            target_total = view.loc[mask_fix, "보정_예상공급량(MJ)"].sum()
                            ideal_pattern = view.loc[mask_fix, "Group_Mean"]
                            if ideal_pattern.sum() > 0:
                                view.loc[mask_fix, "보정_예상공급량(MJ)"] = ideal_pattern * (target_total / ideal_pattern.sum())
            
            st.caption(f"변동량: {mj_to_gj(diff_mj):,.0f} GJ")

    view["예상공급량(GJ)"] = view["예상공급량(MJ)"].apply(mj_to_gj)
    view["보정_예상공급량(GJ)"] = view["보정_예상공급량(MJ)"].apply(mj_to_gj)
    view["Bound_Upper(GJ)"] = view["Bound_Upper"].apply(mj_to_gj)
    view["Bound_Lower(GJ)"] = view["Bound_Lower"].apply(mj_to_gj)

    fig = go.Figure()

    w1 = view[view["구분"] == "평일1(월,금)"].copy()
    w2 = view[view["구분"] == "평일2(화,수,목)"].copy()
    we = view[view["구분"] == "주말/공휴일"].copy()

    fig.add_trace(go.Bar(x=w1["일"], y=w1["예상공급량(GJ)"], name="평일1(월,금)", marker_color="#1F77B4", width=0.8, hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))
    fig.add_trace(go.Bar(x=w2["일"], y=w2["예상공급량(GJ)"], name="평일2(화,수,목)", marker_color="#87CEFA", width=0.8, hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))
    fig.add_trace(go.Bar(x=we["일"], y=we["예상공급량(GJ)"], name="주말/공휴일", marker_color="#D62728", width=0.8, hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))

    if use_calib:
        mask_changed = (abs(view["예상공급량(MJ)"] - view["보정_예상공급량(MJ)"]) > 1)
        if mask_changed.any():
            target_view = view[mask_changed]
            fig.add_trace(go.Bar(
                x=target_view["일"], 
                y=target_view["보정_예상공급량(GJ)"],
                marker_color="rgba(100, 100, 100, 0.6)", 
                name="보정됨(To-Be)",
                width=0.8,
                hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"
            ))

    fig.add_trace(go.Scatter(x=view["일"], y=view["일별비율"], yaxis="y2", name="비율", line=dict(color='#FF8A80', width=2), hovertemplate="%{x}일: %{y:.4f}<extra>%{name}</extra>"))
    fig.add_trace(go.Scatter(x=view["일"], y=view["Bound_Upper(GJ)"], mode='lines', line=dict(width=0), showlegend=False, hoverinfo='skip'))
    fig.add_trace(go.Scatter(x=view["일"], y=view["Bound_Lower(GJ)"], mode='lines', line=dict(width=0), fill='tonexty', fillcolor='rgba(100,100,100,0.45)', name='범위(±10%)', hoverinfo='skip'))
    
    outliers = view[view["is_outlier"]]
    if not outliers.empty:
        fig.add_trace(go.Scatter(x=outliers["일"], y=outliers["예상공급량(GJ)"], mode='markers', marker=dict(color='red', symbol='x', size=10), name='Outlier', hovertemplate="%{x}일: %{y:,.0f} GJ<extra>%{name}</extra>"))

    fig.update_layout(
        title=f"{target_year}년 {target_month}월 공급계획",
        xaxis_title="일",
        yaxis=dict(title="공급량(GJ)"),
        yaxis2=dict(title="비율", overlaying="y", side="right"),
        barmode="overlay", 
        legend=dict(orientation="h", y=1.1)
    )
    
    chart_placeholder.plotly_chart(fig, use_container_width=True)
    
    st.divider()

    st.markdown("### 🧩 1. 일별 공급량 분배 기준")
    st.markdown(
        """
- **주말/공휴일/명절**: **'요일(토/일) + 그 달의 n번째' 기준 평균** (공휴일/명절도 주말 패턴으로 묶음)
- **평일**: '평일1(월,금)' / '평일2(화,수,목)'로 구분  
  기본은 **'요일 + 그 달의 n번째(1째 월요일, 2째 월요일...)' 기준 평균**
- 일부 케이스 데이터가 부족하면 **'요일 평균'으로 보정**
- 마지막에 **일별비율 합계가 1이 되도록 정규화(raw / SUM(raw))**
        """.strip()
    )

    st.markdown("#### 📌 2. 월별 계획량(1~12월) & 연간 총량")
    df_plan_h = make_month_plan_horizontal(df_plan, target_year, plan_col)
    show_table_no_index(format_table_generic(df_plan_h), height=160)

    st.markdown("#### 📋 3. 일별 비율, 예상 공급량 테이블")
    
    total_row = {
        "연": "", "월": "", "일": "", "일자": pd.Timestamp("NaT"), "요일": "합계",
        "weekday_idx": "", "nth_dow": "", "구분": "", "공휴일여부": False,
        "최근N년_평균공급량(MJ)": view["최근N년_평균공급량(MJ)"].sum(),
        "최근N년_총공급량(MJ)": view["최근N년_총공급량(MJ)"].sum(),
        "일별비율": view["일별비율"].sum(),
        "예상공급량(MJ)": view["예상공급량(MJ)"].sum(),
        "보정_예상공급량(MJ)": view["보정_예상공급량(MJ)"].sum(),
    }
    view_with_total = pd.concat([view, pd.DataFrame([total_row])], ignore_index=True)
    view_show = _make_display_table_gj_m3(view_with_total)
    
    if "is_outlier" in view_show.columns:
        view_show["is_outlier"] = view_show["is_outlier"].map({True: "🚨", False: ""})

    view_show = format_table_generic(view_show, percent_cols=["일별비율"])
    show_table_no_index(view_show, height=520)

    st.markdown("#### 🧊 4. 최근 N년 일별 실적 매트릭스")
    if df_mat is not None:
        df_mat_gj = df_mat.map(mj_to_gj)
        fig_hm = go.Figure(
            data=go.Heatmap(
                z=df_mat_gj.values,
                x=[str(c) for c in df_mat_gj.columns],
                y=df_mat_gj.index,
                colorbar_title="공급량(GJ)",
                colorscale="RdBu_r",
            )
        )
        fig_hm.update_layout(
            title=f"실적 매트릭스",
            xaxis=dict(title="연도", type="category"),
            yaxis=dict(title="일", autorange="reversed"),
            margin=dict(l=40, r=40, t=60, b=40),
        )
        st.plotly_chart(fig_hm, use_container_width=False)

    st.markdown("#### 🧾 5. 구분별 비중 요약(평일1/평일2/주말)")
    summary = (
        view.groupby("구분", as_index=False)[["일별비율", "예상공급량(MJ)", "보정_예상공급량(MJ)"]]
        .sum()
        .rename(columns={"일별비율": "일별비율합계"})
    )
    summary["예상공급량(GJ)"] = summary["예상공급량(MJ)"].apply(mj_to_gj).round(0)
    summary["보정_예상공급량(GJ)"] = summary["보정_예상공급량(MJ)"].apply(mj_to_gj).round(0)
    total_row_sum = {
        "구분": "합계",
        "일별비율합계": summary["일별비율합계"].sum(),
        "예상공급량(GJ)": summary["예상공급량(GJ)"].sum(),
        "보정_예상공급량(GJ)": summary["보정_예상공급량(GJ)"].sum(),
    }
    summary = pd.concat([summary, pd.DataFrame([total_row_sum])], ignore_index=True)
    summary_show = summary[["구분", "일별비율합계", "예상공급량(GJ)", "보정_예상공급량(GJ)"]].copy()
    summary_show = format_table_generic(summary_show, percent_cols=["일별비율합계"])
    show_table_no_index(summary_show, height=220)

    st.markdown("#### 💾 6. 데이터 다운로드")
    
    col_down1, col_down2 = st.columns(2)
    
    with col_down1:
        if use_calib:
            st.info("💡 보정된(To-Be) 데이터를 다운로드할 수 있습니다.")
            buffer_tobe = BytesIO()
            dl_src = view_with_total.copy()
            dl_src["As-Is(기존)"] = dl_src["예상공급량(MJ)"].apply(mj_to_gj).round(0)
            dl_src["To-Be(보정)"] = dl_src["보정_예상공급량(MJ)"].apply(mj_to_gj).round(0)
            dl_src["Diff(증감)"] = dl_src["To-Be(보정)"] - dl_src["As-Is(기존)"]
            
            dl_src["Diff(%)"] = dl_src.apply(
                lambda row: (row["Diff(증감)"] / row["As-Is(기존)"] * 100) if row["As-Is(기존)"] != 0 else 0, axis=1
            )

            if "is_outlier" not in dl_src.columns: dl_src["is_outlier"] = ""
            cols_fin = ["일자", "요일", "구분", "As-Is(기존)", "To-Be(보정)", "Diff(증감)", "Diff(%)", "is_outlier"]
            cols_fin = [c for c in cols_fin if c in dl_src.columns]
            
            download_df = dl_src[cols_fin].copy()
            
            with pd.ExcelWriter(buffer_tobe, engine="openpyxl") as writer:
                download_df.to_excel(writer, index=False, sheet_name="To-Be_일별계획")
                
            st.download_button(
                label="📥 To-Be(보정후) 일별계획 다운로드", 
                data=buffer_tobe.getvalue(), 
                file_name=f"{target_year}_{target_month:02d}_ToBe_일별공급계획.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            buffer = BytesIO()
            excel_df = _make_display_table_gj_m3(view_with_total)
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                excel_df.to_excel(writer, index=False, sheet_name="일별계획")
            st.download_button(
                label="📥 일별계획 다운로드 (As-Is)", 
                data=buffer.getvalue(), 
                file_name=f"{target_year}_{target_month:02d}_일별공급계획.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col_down2:
        st.markdown("**🗂️ 연간 전체 계획 다운로드**")
        annual_year = st.selectbox("연간 계획 연도", years_plan, index=years_plan.index(target_year) if target_year in years_plan else 0)
        buffer_year = BytesIO()
        df_year_daily, df_month_summary = _build_year_daily_plan(df_daily, df_plan, int(annual_year), recent_window)
        with pd.ExcelWriter(buffer_year, engine="openpyxl") as writer:
            df_year_daily.to_excel(writer, index=False, sheet_name="연간")
            df_month_summary.to_excel(writer, index=False, sheet_name="월 요약")
            _add_cumulative_status_sheet(writer.book, int(annual_year))
        st.download_button(
            label="📥 연간 계획 다운로드", 
            data=buffer_year.getvalue(), 
            file_name=f"{annual_year}_연간_일별공급계획.xlsx", 
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            key="download_annual_excel"
        )


def main():
    df, _ = load_daily_data()
    mode = st.sidebar.radio("좌측 탭 선택", ("📅 Daily 공급량 분석",), index=0)
    if mode == "📅 Daily 공급량 분석":
        st.title("도시가스 공급량 — 일별계획 예측")
        tab_daily_plan(df_daily=df)

if __name__ == "__main__":
    main()
